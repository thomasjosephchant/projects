import json
import logging
import io
from io import StringIO 
import boto3
import pandas as pd
from openpyxl import load_workbook

logger = logging.getLogger()
logger.setLevel(logging.INFO)

def s3_event_info(event):
    '''
    Extracts the S3 bucket name and file key from a json created by an SNS trigger in response to a file being uploaded.
    
    Parameters: 
        event (json): a json file created from an SNS trigger. This event is created when a file is uploaded to the bucket to which this lambda will be applied. 
    
    Returns: 
        bucket_name (str): The s3 bucket name to which the file is uploaded
        file_key_attribute (str): The file key directing to the location of the file within the s3 bucket.
    '''
    
    # Get the details of the uploaded file from the event
    messagejson = json.loads(event['Records'][0]['Sns']['Message'])
    logging.info("Event has been loaded in.")
    
    # Extract the bucket key
    bucket_name = messagejson['Records'][0]['s3']['bucket']['name']
    if not isinstance(bucket_name, str):
        raise TypeError("The bucket_name grabbed was not a string")
    logging.info(f"The bucket name is: {bucket_name}")
    
    file_key_attribute = messagejson['Records'][0]['s3']['object']['key']
    if not isinstance(file_key_attribute, str):
        raise TypeError("The file_key grabbed was not a string")
    logging.info(f"The file key is: {file_key_attribute}")
    
    return [bucket_name, file_key_attribute]

def file_downloader(bucket_name, file_key_attribute):
    '''
    Downloads a file to the temporary work space
    
    Parameters: 
        bucket_name (str): The s3 bucket name from where the file can be downloaded
        file_key_attribute (str): The file key directing to the location of the file within the s3 bucket.
    
    Returns: 
        body (obj): An io file object containing the data from the file uploaded. A spreadsheet and dataframe can be extracted from this.
    '''
    
    # This function downloads the file into an object from the bucket name and file key
    s3 = boto3.client('s3')
    obj = s3.get_object(Bucket=bucket_name, Key=file_key_attribute)  # download the files to a temp workspace
    body = obj['Body'].read()
    logging.info("Body and obj loaded in")
    return body

def sheet_name_grabber(iofile):
    '''
    Read the bytes file into a spreadsheet and extract the sheet names
    
    Parameters:
        iofile (obj): An io file object containing the data from the file uploaded.
    
    Returns:
        attribute_ex (obj): The loaded workbook object.
        sheet_names_attribute (list): The list of sheet names extracted from the workbook.
    '''
    
    # Read the bytes file into a spreadsheet and extract the sheet names
    attribute_ex = load_workbook(io.BytesIO(iofile), data_only=True)
    logging.info("The workbook has been loaded into attribute_ex")
    # Could add a time out error here if this takes too long
    sheet_names_attribute = attribute_ex.sheetnames
    logging.info(f"The list of sheet names is: {sheet_names_attribute}")
    return attribute_ex, sheet_names_attribute

def tagger(workbook, sheetnamelist):
    '''
    Create a list of tags and a dataframe containing the sheet names and the tags
    
    Parameters:
        workbook (obj): The loaded workbook object.
        sheetnamelist (list): The list of sheet names extracted from the workbook.
    
    Returns:
        df (DataFrame): A dataframe containing the sheet names and tags.
    '''
    
    # Create the list of tags and a dataframe containing the sheet names and the tags
    titles = []
    tags = []

    for sheet_name in sheetnamelist:
        if 'Table' in sheet_name:
            ws = workbook[sheet_name]
            df = pd.DataFrame(ws.values)
            titles.append(df.iloc[0, 2])
            tags.append(str(df.iloc[0, 2]).split('.'))

    # Create a dataframe of the titles and tags        
    df = pd.DataFrame(data=titles, columns=['Table_Name/Kafka_Name'])
    df['Tags'] = tags
    logging.info(f"The list of tags are: {df['Tags']}")
    return df

def uploader(key, df, bucket_name):
    '''
    Upload the dataframe to the location originally extracted from the event
    
    Parameters:
        key (str): The file key directing to the location of the file within the s3 bucket.
        df (DataFrame): The dataframe to be uploaded.
        bucket_name (str): The s3 bucket name to which the file will be uploaded.
    
    Returns:
        message (str): The success message indicating the file has been uploaded.
    '''
    
    file_name = key.replace('.xlsx', '_tagged.csv')
    logging.info(f"The new file name will be: {file_name}")    

    newhome = 'Tags/' + file_name
    logging.info(f"The new home for the file is: {newhome}")    

    # Create the newly updated excel files, replacing the locally saved originals
    csv_buffer = StringIO() 
    if not isinstance(df, pd.DataFrame):
        raise TypeError("df used is not filetype: DataFrame")
    logging.info(f"The bucket name is: {bucket_name}")
    
    df.to_csv(csv_buffer)
    s3_resource = boto3.resource('s3') 
    
    s3_resource.Object(bucket_name, newhome).put(Body=csv_buffer.getvalue()) 
    message = (f"The file has been uploaded to {bucket_name}")
    logging.info(message)
    
    return message

def s3_tagger(event, context):
    '''
    Lambda function to tag the spreadsheet file and upload the tagged file to S3
    
    Parameters:
        event (json): a json file created from an SNS trigger. This event is created when a file is uploaded to the bucket to which this lambda will be applied.
    
    Returns:
        message (str): The success message indicating the file has been uploaded.
    '''
    
    bucket, key = s3_event_info(event)
    body = file_downloader(bucket, key)
    attribute_ex, sheet_names_attribute = sheet_name_grabber(body)
    df = tagger(attribute_ex, sheet_names_attribute)
    message = uploader(key, df, bucket)
    return message